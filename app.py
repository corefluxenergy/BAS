import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

st.set_page_config(page_title="GST Working Papers", layout="wide")

st.title("GST Working Papers Generator (AU BAS)")
st.markdown(
    "Upload your **Commonwealth** and **Wise** CSV files. "
    "Review GST decisions and export BAS-ready working papers."
)

# ------------------------
# Styling
# ------------------------
st.markdown(
    """
    <style>
    .metric-box {
        background-color: #1f2937;
        padding: 20px;
        border-radius: 10px;
        text-align: center;
        color: white;
    }
    .metric-label {
        font-size: 14px;
        opacity: 0.8;
    }
    .metric-value {
        font-size: 28px;
        font-weight: 700;
        margin-top: 5px;
    }
    div.stDownloadButton > button {
        width: 100%;
        height: 3em;
        font-size: 18px;
        font-weight: 600;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ------------------------
# Upload section
# ------------------------
col1, col2 = st.columns(2)

with col1:
    cba_file = st.file_uploader("Upload Commonwealth CSV", type="csv")

with col2:
    wise_file = st.file_uploader("Upload Wise CSV", type="csv")

if not cba_file or not wise_file:
    st.info("Please upload both CSV files to continue.")
    st.stop()

# ------------------------
# Load & normalise data
# ------------------------

# Commonwealth
cba = pd.read_csv(cba_file, header=None)
cba.columns = ["Date", "Amount", "Description", "Balance"]
cba["Date"] = pd.to_datetime(cba["Date"], errors="coerce", dayfirst=True)
cba["Account"] = "Commonwealth"
cba["Direction"] = cba["Amount"].apply(lambda x: "IN" if x > 0 else "OUT")
cba["Amount"] = pd.to_numeric(cba["Amount"], errors="coerce").fillna(0)

# Wise
wise = pd.read_csv(wise_file)
wise["Date"] = pd.to_datetime(wise["Finished on"], errors="coerce")
wise["Account"] = "Wise"
wise["Description"] = wise["Source name"]
wise["Direction"] = wise["Direction"]
wise["Amount"] = pd.to_numeric(
    wise["Target amount (after fees)"], errors="coerce"
).fillna(0)

ledger = pd.concat(
    [
        cba[["Date", "Account", "Description", "Direction", "Amount"]],
        wise[["Date", "Account", "Description", "Direction", "Amount"]],
    ],
    ignore_index=True,
)

# ------------------------
# Determine BAS Quarter
# ------------------------
def month_to_quarter(m):
    if m <= 3:
        return "Q1"
    if m <= 6:
        return "Q2"
    if m <= 9:
        return "Q3"
    return "Q4"

ledger["Quarter"] = ledger["Date"].dt.month.apply(month_to_quarter)
bas_quarter = ledger["Quarter"].mode().iloc[0]
sheet_name = f"{bas_quarter} – BAS GST"

# ------------------------
# Classification logic
# ------------------------
def classify(row):
    desc = str(row["Description"]).lower()

    if row["Direction"] == "IN":
        return "Income", "NO", "Income received"

    if "transfer" in desc:
        return "Transfer", "NO", "Internal transfer"

    if any(k in desc for k in ["fee", "fx", "asic", "ato", "bpay", "tax"]):
        return "Fee", "NO", "GST-free fee or government charge"

    return "Expense", "YES", "Australian business expense – GST assumed"


ledger[["Transaction Type", "GST Claimable", "System Reason"]] = ledger.apply(
    lambda r: pd.Series(classify(r)), axis=1
)

ledger["Gross Amount"] = ledger["Amount"].abs().round(2)
ledger["GST Amount"] = 0.0
ledger["Net (ex GST)"] = ledger["Gross Amount"]
ledger["Comment"] = ""

# ------------------------
# Editable ledger
# ------------------------
st.subheader("Ledger")

edited_df = st.data_editor(
    ledger.drop(columns=["Quarter"]),
    use_container_width=True,
    num_rows="fixed",
    column_config={
        "GST Claimable": st.column_config.SelectboxColumn(
            "GST Claimable", options=["YES", "NO"]
        ),
        "Comment": st.column_config.TextColumn("Comment"),
    },
)

# ------------------------
# GST SUMMARY – BIG HTML CARDS (FIXED)
# ------------------------
income_gross = edited_df.loc[
    edited_df["Transaction Type"] == "Income", "Gross Amount"
].sum()

gst_on_sales = income_gross / 11
gst_claimable_gross = edited_df.loc[
    edited_df["GST Claimable"] == "YES", "Gross Amount"
].sum()
gst_on_purchases = gst_claimable_gross / 11
net_gst = gst_on_sales - gst_on_purchases

st.subheader("GST Summary")

c1, c2, c3, c4, c5 = st.columns(5)

metrics = [
    ("G1 – Total sales", income_gross),
    ("1A – GST on sales", gst_on_sales),
    ("GST-claimable expenses", gst_claimable_gross),
    ("1B – GST on purchases", gst_on_purchases),
    ("Net GST payable", net_gst),
]

for col, (label, value) in zip([c1, c2, c3, c4, c5], metrics):
    col.markdown(
        f"""
        <div class="metric-box">
            <div class="metric-label">{label}</div>
            <div class="metric-value">${value:,.2f}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ------------------------
# Excel export
# ------------------------
def export_excel(df):
    output = BytesIO()
    currency_fmt = '$#,##0.00'

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        last_row = len(df) + 1

        amount_col = get_column_letter(df.columns.get_loc("Amount") + 1)
        gross_col = get_column_letter(df.columns.get_loc("Gross Amount") + 1)
        gst_flag_col = get_column_letter(df.columns.get_loc("GST Claimable") + 1)
        gst_amt_col = get_column_letter(df.columns.get_loc("GST Amount") + 1)
        net_col = get_column_letter(df.columns.get_loc("Net (ex GST)") + 1)
        type_col = get_column_letter(df.columns.get_loc("Transaction Type") + 1)

        # GST dropdown
        dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{gst_flag_col}2:{gst_flag_col}{last_row}")

        # Row formulas + currency formatting
        for r in range(2, last_row + 1):
            ws[f"{gst_amt_col}{r}"] = f'=IF({gst_flag_col}{r}="YES",{gross_col}{r}/11,0)'
            ws[f"{net_col}{r}"] = f'={gross_col}{r}-{gst_amt_col}{r}'

            for col in [amount_col, gross_col, gst_amt_col, net_col]:
                ws[f"{col}{r}"].number_format = currency_fmt

        # Summary
        s = last_row + 3
        ws[f"A{s}"] = "GST SUMMARY (AUTO-CALCULATED)"
        ws[f"A{s}"].font = Font(bold=True)

        ws[f"A{s+2}"] = "G1 – Total sales (incl GST)"
        ws[f"B{s+2}"] = f'=SUMIF({type_col}2:{type_col}{last_row},"Income",{gross_col}2:{gross_col}{last_row})'

        ws[f"A{s+3}"] = "1A – GST on sales"
        ws[f"B{s+3}"] = f"=B{s+2}/11"

        ws[f"A{s+4}"] = "GST-claimable expenses (gross)"
        ws[f"B{s+4}"] = f'=SUMIF({gst_flag_col}2:{gst_flag_col}{last_row},"YES",{gross_col}2:{gross_col}{last_row})'

        ws[f"A{s+5}"] = "1B – GST on purchases"
        ws[f"B{s+5}"] = f'=SUMIF({gst_flag_col}2:{gst_flag_col}{last_row},"YES",{gst_amt_col}2:{gst_amt_col}{last_row})'

        ws[f"A{s+6}"] = "Net GST payable"
        ws[f"B{s+6}"] = f"=B{s+3}-B{s+5}"

        for r in range(s + 2, s + 7):
            ws[f"B{r}"].number_format = currency_fmt

        # Force-safe widths for numeric columns
        ws.column_dimensions[amount_col].width = 18
        ws.column_dimensions[gross_col].width = 18
        ws.column_dimensions[gst_amt_col].width = 18
        ws.column_dimensions[net_col].width = 18

        # Autosize remaining columns
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            if ws.column_dimensions[letter].width:
                continue
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[letter].width = max_len + 2

    return output.getvalue()

# ------------------------
# Export
# ------------------------
st.divider()
st.subheader("Export")

st.download_button(
    label="⬇️ Export Excel (BAS GST)",
    data=export_excel(edited_df),
    file_name=f"{bas_quarter}_BAS_GST.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
